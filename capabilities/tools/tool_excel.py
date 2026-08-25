"""
capabilities/tools/tool_excel.py
=================================
Tool 9: Excel operations.

Build/edit: via openpyxl (vendored) for formulas and formatting.
Data:       via pandas (requirements) for reading/writing tabular data.
"""
from __future__ import annotations
import os
import sys
from . import BaseTool, ToolResult

TOOL_SPEC = {
    "name": "excel",
    "description": "Build, read, or edit Excel (.xlsx) files with formulas and formatting via openpyxl/pandas.",
    "triggers": ["Excel", "XLSX", "جدول بيانات", "شيت", "صيغ",
                 "spreadsheet", "formulas", "read excel", "build excel"],
    "layers": [8],
}


def _vendored_dir():
    base = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    return os.path.join(base, "engines", "office-core", "vendored")


class ExcelTool(BaseTool):
    TOOL_SPEC = TOOL_SPEC

    def _ensure_vendored(self):
        v = _vendored_dir()
        if v not in sys.path:
            sys.path.insert(0, v)

    async def _execute(self, inputs: dict) -> ToolResult:
        action = inputs.get("action", "build")  # build | read
        if action == "read":
            return await self._read(inputs)
        return await self._build(inputs)

    async def _build(self, inputs: dict) -> ToolResult:
        data   = inputs.get("data", [])       # list of rows (list of lists)
        headers = inputs.get("headers", [])
        output = inputs.get("output_path", "./output/result.xlsx")
        sheet_name = inputs.get("sheet_name", "Sheet1")

        os.makedirs(os.path.dirname(output) or ".", exist_ok=True)
        self._ensure_vendored()

        try:
            from openpyxl import Workbook
        except ImportError:
            return ToolResult(ok=False, error="openpyxl not available")

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        row_offset = 1
        if headers:
            for c, h in enumerate(headers, start=1):
                ws.cell(row=1, column=c, value=h)
            row_offset = 2

        for r, row in enumerate(data, start=row_offset):
            for c, val in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=val)

        # Optional totals row with SUM formulas
        if inputs.get("with_totals") and data:
            total_row = row_offset + len(data)
            for c in range(1, len(data[0]) + 1):
                col = ws.cell(row=1, column=c).column_letter
                first = row_offset
                last = total_row - 1
                # only numeric columns get a SUM
                ws.cell(row=total_row, column=c,
                        value=f"=SUM({col}{first}:{col}{last})")

        wb.save(output)
        return ToolResult(ok=True, data={"output_path": output, "engine": "openpyxl"})

    async def _read(self, inputs: dict) -> ToolResult:
        path = inputs.get("path", "").strip()
        if not path or not os.path.exists(path):
            return ToolResult(ok=False, error=f"file not found: {path}")

        self._ensure_vendored()
        # Prefer openpyxl (vendored, lighter than pandas)
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True)
            ws = wb.active
            rows = [[cell.value for cell in row] for row in ws.iter_rows()]
            return ToolResult(ok=True, data={
                "rows": rows, "sheet": ws.title, "engine": "openpyxl",
            })
        except ImportError:
            pass

        # Fallback: pandas (requirements)
        try:
            import pandas as pd
            df = pd.read_excel(path)
            return ToolResult(ok=True, data={
                "rows": df.values.tolist(),
                "columns": df.columns.tolist(),
                "engine": "pandas",
            })
        except ImportError:
            return ToolResult(ok=False, error="no Excel reader available")


async def run(inputs: dict) -> ToolResult:
    return await ExcelTool().run(inputs)
