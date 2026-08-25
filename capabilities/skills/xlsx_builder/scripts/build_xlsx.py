"""
build_xlsx.py — build a professional Excel workbook (working script)
====================================================================
Direction-aware:
  - Arabic (lang="ar")  -> sheet RTL (sheetView rightToLeft), headers right.
  - English (lang="en") -> sheet LTR.

Features: styled header row (navy fill, white bold), borders, auto width,
optional SUM totals row, number formatting.

Requires: pip install openpyxl
"""
from __future__ import annotations
import argparse
import json

NAVY = "1B2A4A"
GOLD = "C8A04A"
WHITE = "FFFFFF"


def build_xlsx(data, output_path, headers=None, lang="ar",
               sheet_name=None, with_totals=False, number_cols=None):
    """Build a formatted workbook. lang 'ar' sets the sheet right-to-left."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    rtl = (lang == "ar")
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name or ("البيانات" if rtl else "Data")
    ws.sheet_view.rightToLeft = rtl   # <-- the key RTL switch

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor=NAVY)
    header_font = Font(bold=True, color=WHITE, size=12)
    align = Alignment(horizontal=("right" if rtl else "left"), vertical="center")
    center = Alignment(horizontal="center", vertical="center")

    row_offset = 1
    if headers:
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        row_offset = 2

    for r, row in enumerate(data, start=row_offset):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = align
            cell.border = border

    # totals row with SUM formulas over numeric columns
    if with_totals and data:
        total_row = row_offset + len(data)
        ncols = len(data[0])
        for c in range(1, ncols + 1):
            col_letter = ws.cell(row=1, column=c).column_letter
            if number_cols is None or c in number_cols:
                # only sum columns that look numeric
                sample = data[0][c-1]
                if isinstance(sample, (int, float)):
                    cell = ws.cell(row=total_row, column=c,
                                   value=f"=SUM({col_letter}{row_offset}:{col_letter}{total_row-1})")
                    cell.font = Font(bold=True, color=NAVY)
                    cell.border = border
                    cell.alignment = align
            if c == 1:
                lbl = ws.cell(row=total_row, column=1,
                              value=("الإجمالي" if rtl else "Total"))
                lbl.font = Font(bold=True, color=NAVY)

    # auto column width
    for col in ws.columns:
        width = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 4, 50)

    wb.save(output_path)
    return output_path


def _main():
    p = argparse.ArgumentParser(description="Build a professional Excel workbook")
    p.add_argument("--json", required=True)
    p.add_argument("--output", default="data.xlsx")
    p.add_argument("--lang", default="ar", choices=["ar", "en"])
    args = p.parse_args()
    with open(args.json, encoding="utf-8") as f:
        d = json.load(f)
    path = build_xlsx(d.get("data", []), args.output, headers=d.get("headers"),
                      lang=args.lang, with_totals=d.get("with_totals", False))
    print(f"Created: {path}")


if __name__ == "__main__":
    _main()
